from openpyxl import load_workbook
from copy import copy

INPUT_FILE = "Dabar Details.xlsx"
OUTPUT_FILE = "Dabar Details English.xlsx"

# Shivaji legacy text -> English
translations = {
    "idnaaMk": "Date",
    "gaaDI naM": "Vehicle No.",
    "Kopa": "Trips",
    "ekuNa Kopa": "Total Trips",
    "ekuNa Dbar Kopa": "Total Dumper Trips",
    "A.k`": "Sr. No.",
    "A.k": "Sr. No.",
}

wb = load_workbook(INPUT_FILE)

for ws in wb.worksheets:

    for row in ws.iter_rows():
        for cell in row:

            if not isinstance(cell.value, str):
                continue

            text = cell.value

            # Exact matches first
            if text in translations:
                cell.value = translations[text]

            # Longer heading containing multiple Shivaji words
            elif "k`oSarvar Aalaolao" in text:
                text = text.replace(
                    "k`oSarvar Aalaolao",
                    "Received at Crusher"
                )

                text = text.replace(
                    "ekuNa Dbar Kopa",
                    "Total Dumper Trips"
                )

                text = text.replace(
                    "pya-t",
                    "up to"
                )

                cell.value = text

            # Total heading
            elif "ekuNa Dbar Kopa" in text:
                cell.value = text.replace(
                    "ekuNa Dbar Kopa",
                    "Total Dumper Trips"
                )

            # Change font for converted text
            if isinstance(cell.value, str):
                if cell.value in translations.values():
                    new_font = copy(cell.font)
                    new_font.name = "Noto Sans"
                    cell.font = new_font

# Save new workbook
wb.save(OUTPUT_FILE)

print("Conversion completed!")
print(f"Output file: {OUTPUT_FILE}")
