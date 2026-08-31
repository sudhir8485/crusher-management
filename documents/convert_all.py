from openpyxl import load_workbook
from copy import copy
from pathlib import Path
import re

# Folder containing your Excel files
INPUT_FOLDER = Path(".")

# Output folder
OUTPUT_FOLDER = Path("converted")
OUTPUT_FOLDER.mkdir(exist_ok=True)

# Shivaji legacy text -> English
TRANSLATIONS = {
    "idnaaMk": "Date",
    "gaaDI naM": "Vehicle No.",
    "Kopa": "Trips",
    "ekuNa Kopa": "Total Trips",
    "ekuNa Dbar Kopa": "Total Dumper Trips",
    "A.k`": "Sr. No.",
    "A.k": "Sr. No.",
}

# Exact text replacements inside longer headings
REPLACEMENTS = {
    "k`oSarvar Aalaolao": "Received at Crusher",
    "ekuNa Dbar Kopa": "Total Dumper Trips",
    "ekuNa Kopa": "Total Trips",
    "gaaDI naM": "Vehicle No.",
    "idnaaMk": "Date",
    "pya-t": "up to",
    "A.k`": "Sr. No.",
}

def is_date_or_number(text):
    """
    Do not translate dates, numbers, vehicle numbers, etc.
    """
    text = text.strip()

    # Dates / numbers / punctuation
    if re.fullmatch(r"[0-9.,:/\-\s]+", text):
        return True

    return False


def convert_text(text):
    """
    Convert known Shivaji text to English.
    """

    if not isinstance(text, str):
        return text, False

    original = text

    # Ignore empty strings
    if not text.strip():
        return text, False

    # Don't touch dates/numbers
    if is_date_or_number(text):
        return text, False

    # Exact translation
    if text.strip() in TRANSLATIONS:
        return TRANSLATIONS[text.strip()], True

    # Replace known Shivaji phrases
    converted = text

    for old, new in REPLACEMENTS.items():
        converted = converted.replace(old, new)

    # If something changed, return it
    if converted != original:
        return converted, True

    return text, False


def process_workbook(input_file, output_file):

    print(f"\nProcessing: {input_file.name}")

    try:
        wb = load_workbook(input_file)

        total_converted = 0

        for ws in wb.worksheets:

            for row in ws.iter_rows():

                for cell in row:

                    if not isinstance(cell.value, str):
                        continue

                    new_value, changed = convert_text(cell.value)

                    if changed:

                        cell.value = new_value

                        # Keep all existing formatting,
                        # but use a normal Unicode-friendly font.
                        new_font = copy(cell.font)
                        new_font.name = "Noto Sans Devanagari"
                        cell.font = new_font

                        total_converted += 1

        wb.save(output_file)

        print(f"  Sheets processed : {len(wb.worksheets)}")
        print(f"  Text converted   : {total_converted}")
        print(f"  Saved to         : {output_file}")

    except Exception as e:

        print(f"  ERROR: {e}")


# Process every XLSX file
files = sorted(INPUT_FOLDER.glob("*.xlsx"))

# Don't process temporary Excel files
files = [
    f for f in files
    if not f.name.startswith("~$")
]

if not files:
    print("No .xlsx files found.")

else:

    print(f"Found {len(files)} Excel files.")

    for input_file in files:

        output_file = OUTPUT_FOLDER / input_file.name

        process_workbook(
            input_file,
            output_file
        )

    print("\n================================")
    print("ALL FILES PROCESSED")
    print("================================")
    print(f"Output folder: {OUTPUT_FOLDER.resolve()}")
